import re
import json
import openpyxl
import ast
import string
from xlsxwriter import Workbook
from typing import List, Dict, Any, Tuple

INPUT_FILE_PATH_HAN_NOM = "data/clean/TongTapVanHocVietNam06_clean.hannom.json"
INPUT_FILE_PATH_QUOC_NGU = "data/clean/TongTapVanHocVietNam06_clean.quocngu.json"
OUTPUT_FILE_PATH = "result/result.xlsx"
QUOCNGU_SINONOM_DIC_FILE = "resources/QuocNgu_SinoNom_Dic.xlsx"
SINONOM_SIMILAR_DIC_FILE = "resources/SinoNom_similar_Dic.xlsx"

COL_IMAGE_NAME = 0
COL_ID = 1
COL_IMAGE_BOX = 2
COL_SINONOM_OCR = 3
COL_CHU_QUOC_NGU = 4
COL_NGHIA_THUAN_VIET = 5


# Đọc dữ liệu hán nôm ocr từ file JSON
def read_input_file_han_nom(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)
        return data
    
def read_input_file_quoc_ngu(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)
        return data

# Tạo workbook để ghi vào file xlsx
def create_workbook(filename):
    workbook = Workbook(filename)
    worksheet = workbook.add_worksheet("SinoNom Data")
    return workbook, worksheet

# Khởi tạo header của từng cột
def set_headers():
    headers = ["Image_name", "ID", "Image Box", "SinoNom OCR", "Chữ Quốc Ngữ", "Nghĩa Thuần Việt"]
    bold_format = workbook.add_format({'bold': True})
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, bold_format)

# Tải dữ liệu từ điển (SinoNom_similar_Dic & QuocNgu_SinoNom_Dic) vào  list 
def load_dictionary(file_path, key1, key2):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    return [
        {key1: row[0], key2: row[1]}
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ]

# Xử lý định dạng chữ quốc ngữ 
def preprocess_quocngu_text(text):
    text = text.lower()
    text = text.translate(str.maketrans('', '', string.punctuation))
    return text.split()

cache_dp = None

# Tính ma trận Levenshtein
def compute_levenshtein_matrix(str1, str2):
    m, n = len(str1), len(str2)
    global cache_dp

    # Nếu cache_dp chưa được khởi tạo
    if not cache_dp:
        dp = [[0] * (n + 1) for _ in range(m + 1)]

        for i in range(m + 1):
            dp[i][0] = i
        for j in range(n + 1):
            dp[0][j] = j
        
        for i in range(1, m + 1):
            for j in range(1, n + 1):
                # val=2 nếu kết quả giao bằng 0
                # val=0 nết kết quả giao >= 1
                val = 2 if intersect_characters(str1[i-1], str2[j-1]) == 0 else 0
                dp[i][j] = min( dp[i-1][j]+1,           # deletion
                                dp[i][j-1]+1,           # insertion
                                dp[i-1][j-1] + val)     # subtitution
                
        cache_dp = dp
        return cache_dp
    
    # Nếu cache_dp đã được khởi tạo
    if len(cache_dp[0]) <= n:
        old_n = len(cache_dp[0]) - 1
        new_dp = [row[:] for row in cache_dp]
        for i in range(m + 1):
            for j in range(old_n + 1, n + 1):
                new_dp[i].append(0)
                if i == 0:
                    new_dp[0][j] = j
                else:
                    val = 2 if intersect_characters(str1[i-1], str2[j-1]) == 0 else 0
                    new_dp[i][j] = min( new_dp[i-1][j] + 1,
                                        new_dp[i][j-1] + 1,
                                        new_dp[i-1][j-1] + val)
        
        cache_dp = new_dp
        return new_dp
    
    return cache_dp

# Xóa cache của ma trận Levenshtein
def clear_cache_dp():
    global cache_dp
    cache_dp = None

# Các biến cache cho việc lưu trữ kết quả của các hàm tìm kiếm dữ liệu
similar_cache = {}
sino_cache = {}
intersect_cache = {}

# Tìm các từ đồng nghĩa của chữ hán nôm
def find_similar(sinonom_word, data):
    global similar_cache
    if sinonom_word in similar_cache:
        return similar_cache[sinonom_word]
    
    for entry in data:
        if entry['input'] == sinonom_word:
            similar_words = entry['similar']
            if isinstance(similar_words, str):
                similar_words = ast.literal_eval(similar_words)
                similar_words.insert(0, sinonom_word)
            similar_cache[sinonom_word] = set(similar_words)
            return similar_cache[sinonom_word]
        
    similar_cache[sinonom_word] = set()
    return similar_cache[sinonom_word]

# Tìm các chữ hán nôm tương ứng với chữ quốc ngữ
def find_sino(quocngu_word, data):
    global sino_cache
    if quocngu_word in sino_cache:
        return sino_cache[quocngu_word]
    
    sino_cache[quocngu_word] = {entry['sinonom'] for entry in data if entry['quocngu'] == quocngu_word}
    return sino_cache[quocngu_word]

# Tìm số lượng ký tự giao nhau giữa 2 từ
def intersect_characters(sinonom_word, quocngu_word):
    global intersect_cache
    key = (sinonom_word, quocngu_word)
    if key in intersect_cache:
        return intersect_cache[key]
    
    S1 = find_similar(sinonom_word, sino_similar_dic)
    S2 = find_sino(quocngu_word, qn_sino_dic)
    intersect_len = len(S1 & S2)
    result = intersect_len if S1 and S2 else 0
    intersect_cache[key] = result
    return result

# Xóa cache của các hàm tìm kiếm dữ liệu
def clear_caches():
    global similar_cache, sino_cache, intersect_cache
    similar_cache.clear()
    sino_cache.clear()
    intersect_cache.clear()

# Truy vết đường đi của MED levenshtein và tô màu
def backtrace_levenshtein(dp, str1, str2):
    i, j = len(str1), len(str2)
    red = workbook.add_format({'color': 'red'})
    blue = workbook.add_format({'color': 'blue'})
    black = workbook.add_format({'color': 'black'})
    format_pairs = []
    color = {}

    # truy vết đường đi
    while i > 0 and j > 0:
        current_char1 = str1[i - 1]
        current_char2 = str2[j - 1]
        intersection_len = intersect_characters(current_char1, current_char2)

        if dp[i][j] == dp[i - 1][j - 1] + (2 if intersection_len == 0 else 0):
            if intersection_len == 0:
                color[current_char1] = red
            elif intersection_len >= 1:
                color[current_char1] = black
            # else:
            #     color[current_char1] = black
            i -= 1
            j -= 1
        elif dp[i][j] == dp[i - 1][j] + 1:
            color[str1[i - 1]] = red
            i -= 1
        else:
            j -= 1

    while i > 0:
        color[str1[i - 1]] = red
        i -= 1

    # xây dựng format_pairs để tô màu
    for char in str1:
        char_color = color.get(char, red)
        format_pairs.extend((char_color, char))

    format_pairs.extend((black, ' '))
    return format_pairs

# Truy vết đường đi của MED levenshtein đánh dấu
def backtrace_lenven_other(dp, str1, str2):
    i, j = len(str1), len(str2)
    red = workbook.add_format({'color': 'red'})
    blue = workbook.add_format({'color': 'blue'})
    black = workbook.add_format({'color': 'black'})
    format_pairs = []
    color = {}
    results = []
    count_correct = 0
    while i > 0 and j > 0:
        current_char1 = str1[i - 1]
        current_char2 = str2[j - 1]
        intersection_len = intersect_characters(current_char1, current_char2)

        if dp[i][j] == dp[i - 1][j - 1] + (2 if intersection_len == 0 else 0):
            if intersection_len == 0:
                color[current_char1] = red
                results.append({"han_nom": current_char1, "correct": False, "quoc_ngu": None})
            elif intersection_len > 1:
                color[current_char1] = blue
                results.append({"han_nom": current_char1,"correct": True, "quoc_ngu": current_char2})
                count_correct += 1
            else:
                results.append({"han_nom": current_char1, "correct": True, "quoc_ngu": current_char2})
                color[current_char1] = black
                count_correct += 1
            i -= 1
            j -= 1
        elif dp[i][j] == dp[i - 1][j] + 1:
            results.append({"han_nom": current_char1, "correct": False, "quoc_ngu": None})
            color[str1[i - 1]] = red
            i -= 1
        else:
            j -= 1
    
    while i > 0:
        results.append({"han_nom": str1[i - 1], "correct": False, "quoc_ngu": None})
        color[str1[i - 1]] = red
        i -= 1

    # xây dựng format_pairs để tô màu
    for char in str1:
        char_color = color.get(char, red)
        format_pairs.extend((char_color, char))

    format_pairs.extend((black, ' '))
    results.reverse()

    return format_pairs, results, count_correct/len(str1)  

# Tìm các cặp từ tương ứng giữa chữ hán nôm và chữ quốc ngữ
def alignments(hannom_datas, quocngu_datas):
    alignments_result = []
    index_poem = 0
    index_line = 0
    start_index = 0

    for item in hannom_datas:
        file_name = item["file_name"]
        data = item["data"]
        result = []
        for text_line in data:
            if index_poem >= len(quocngu_datas):
                break
                
            hannom_text = text_line["text"].strip()
            position = text_line["position"]
            
            # Skip empty lines in Quốc Ngữ data
            while (index_poem < len(quocngu_datas) and 
                    index_line < len(quocngu_datas[index_poem]["poem_content"]) and 
                    not quocngu_datas[index_poem]["poem_content"][index_line]["phien_am"].strip()):
                index_line += 1
            
            # Move to next poem if we've reached the end of current poem
            if (index_poem >= len(quocngu_datas) or 
                index_line >= len(quocngu_datas[index_poem]["poem_content"])):
                index_poem += 1
                index_line = 0
                start_index = 0
                continue
            
            current_poem = quocngu_datas[index_poem]["poem_content"]
            quocngu_text_raw = current_poem[index_line]["phien_am"].strip()
            quocngu_text = preprocess_quocngu_text(quocngu_text_raw)
            quocngu_text_len = len(quocngu_text)
            
            # Check if text line is long (based on position)
            is_long_line = position[3][1] - position[0][1] >= 80
            is_short_hannom = len(hannom_text) <= quocngu_text_len - 3
            
            if is_long_line and is_short_hannom:
                # Find best matching segment
                best_index = start_index
                best_accuracy = 0
                
                search_start = max(start_index, len(hannom_text) - 3)
                search_end = min(len(hannom_text) + 4 + start_index, quocngu_text_len + 1)
                
                for i in range(search_start, search_end):
                    dp = compute_levenshtein_matrix(hannom_text, quocngu_text[start_index:i])
                    _, _, accuracy = backtrace_lenven_other(dp, hannom_text, quocngu_text[start_index:i])
                    
                    if accuracy > best_accuracy:
                        best_accuracy = accuracy
                        best_index = i
                
                # Add aligned text pair to results
                result.append({
                    "hannom_text": hannom_text,
                    "quocngu_text": " ".join(quocngu_text[start_index:best_index]),
                    "dich_nghia": current_poem[index_line]["dich_nghia"] if start_index == 0 else None,
                    "position": position
                })
                
                clear_cache_dp()
                
                # Update indices
                start_index = best_index
                if best_index >= quocngu_text_len - 2:
                    index_line += 1
                    start_index = 0
            else:
                # Add complete line alignment
                result.append({
                    "hannom_text": hannom_text,
                    "quocngu_text": " ".join(quocngu_text),
                    "dich_nghia": current_poem[index_line]["dich_nghia"],
                    "position": position
                })
                index_line += 1
            
            # Move to next poem if we've finished current one
            if index_line >= len(current_poem):
                index_poem += 1
                index_line = 0
                start_index = 0

        alignments_result.append({
            "file_name": file_name,
            "data": result
        })
    
    return alignments_result
    


def process_data():
    hannom_datas = read_input_file_han_nom(INPUT_FILE_PATH_HAN_NOM)
    quocngu_datas = read_input_file_quoc_ngu(INPUT_FILE_PATH_QUOC_NGU)
    results = alignments(hannom_datas, quocngu_datas)
    row = 1
    for result in results:
        print(f"[+] Processing {result['file_name']}...")
        cnt = 1
        for item in result["data"]:
            id = f"{result['file_name'].replace("_page",".").replace("png","")}{cnt:03}"
            dp = compute_levenshtein_matrix(item["hannom_text"], preprocess_quocngu_text(item["quocngu_text"]))
            format_pairs = backtrace_levenshtein(dp, item["hannom_text"], preprocess_quocngu_text(item["quocngu_text"]))
            worksheet.write(row, COL_ID, id)
            worksheet.write(row, COL_IMAGE_NAME, result['file_name'])
            worksheet.write(row, COL_IMAGE_BOX, str(item["position"]))
            worksheet.write_rich_string(row, COL_SINONOM_OCR, *format_pairs)
            worksheet.write(row, COL_CHU_QUOC_NGU, item["quocngu_text"])
            worksheet.write(row, COL_NGHIA_THUAN_VIET, item["dich_nghia"])
            clear_cache_dp()
            cnt += 1
            row += 1

    workbook.close()

    print("[+] Done!")


def main():
    global workbook, worksheet, qn_sino_dic, sino_similar_dic
    workbook, worksheet = create_workbook(OUTPUT_FILE_PATH)
    set_headers()
    qn_sino_dic = load_dictionary(QUOCNGU_SINONOM_DIC_FILE, 'quocngu', 'sinonom')
    sino_similar_dic = load_dictionary(SINONOM_SIMILAR_DIC_FILE, 'input', 'similar')
    process_data()

if __name__ == "__main__":
    main()